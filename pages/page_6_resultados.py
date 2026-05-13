# coding: utf-8
"""
Página 6: Resultados Finais

Exibe os resultados finais da análise de coastdown e permite exportação.
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
import os
import sys

# Adiciona o diretório raiz ao path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from data.exporters import gerar_excel


def _get(pair, *keys, default=0):
    """Busca o primeiro valor válido dentre as chaves fornecidas."""
    for key in keys:
        val = pair.get(key)
        if val is not None and isinstance(val, (int, float)):
            return val
    return default


def normalize_selected_pairs():
    """Resolve os pares finais para uma lista de dicts, aceitando estados antigos."""
    selected = st.session_state.get("pares_finais_selecionados", [])
    calculated = st.session_state.get("calculated_pairs", {})
    normalized = []
    ignored = 0

    if isinstance(selected, dict):
        selected_items = selected.values()
    elif isinstance(selected, str):
        selected_items = [selected]
    else:
        try:
            selected_items = list(selected)
        except TypeError:
            selected_items = []
            ignored += 1

    for item in selected_items:
        if isinstance(item, dict):
            normalized.append(item)
        elif isinstance(item, str):
            pair = calculated.get(item) if isinstance(calculated, dict) else None
            if isinstance(pair, dict):
                normalized.append(pair)
            else:
                ignored += 1
        else:
            ignored += 1

    return normalized, ignored


def render(t):
    """Renderiza a página de resultados finais."""
    
    st.header(t("page_final_results"))
    
    # Verifica se há resultados finais
    if not st.session_state.final_results:
        st.warning(t("error_select_pairs"))
        return
    
    results = st.session_state.final_results
    selected_pairs, ignored_pairs = normalize_selected_pairs()
    
    # ===== RESUMO DOS RESULTADOS =====
    st.subheader(f"📊 {t('summary')}")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            t("corrected_f0"),
            f"{results.get('mean_f0', 0):.4f} N",
            delta=f"CV: {results.get('cv_f0', 0):.2f}%"
        )
    
    with col2:
        st.metric(
            t("corrected_f2"),
            f"{results.get('mean_f2', 0):.6f}",
            delta=f"CV: {results.get('cv_f2', 0):.2f}%"
        )
    
    with col3:
        st.metric(
            t("total_energy"),
            f"{results.get('mean_energy', 0):.4f} MJ/km"
        )
    
    st.markdown("---")
    
    # ===== INFORMAÇÕES DO VEÍCULO =====
    st.subheader(f"🚗 {t('vehicle_information')}")
    
    vehicle_info = results.get("vehicle_info", {})
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.write(f"**{t('vehicle_model')}:** {vehicle_info.get('model', 'N/A')}")
        st.write(f"**{t('test_date')}:** {vehicle_info.get('test_date', 'N/A')}")
    
    with col2:
        st.write(f"**{t('total_mass')}:** {results.get('total_mass', 0):.1f} kg")
        st.write(f"**{t('effective_mass')}:** {vehicle_info.get('effective_mass', 0):.1f} kg")
    
    with col3:
        st.write(f"**{t('frontal_area')}:** {vehicle_info.get('frontal_area', 0):.2f} m²")
        st.write(f"**{t('selected_pairs')}:** {results.get('num_pairs', 0)}")
    
    st.markdown("---")
    
    # ===== TABELA DE PARES SELECIONADOS =====
    st.subheader(f"📋 {t('selected_pairs')}")
    
    if ignored_pairs:
        st.warning(f"⚠️ {ignored_pairs} par(es) selecionado(s) não foram encontrados e foram ignorados.")

    if selected_pairs:
        pairs_data = []
        for pair in selected_pairs:
            f0 = _get(pair, "f0_corr", "mean_f0_corrected", "f0corr_mean", "f0_mean")
            f2 = _get(pair, "f2_corr", "mean_f2_corrected", "f2corr_mean", "f2_mean")
            energy = _get(pair, "energy", "mean_energy_corrected")
            temp = pair.get("temp", "N/A")
            press = pair.get("press", "N/A")
            
            pair_id = pair.get("pair_id", "N/A")
            if pair_id == "N/A":
                run_ida = pair.get("run_ida", "?")
                run_volta = pair.get("run_volta", "?")
                pair_id = f"{run_ida}/{run_volta}"
            
            pairs_data.append({
                t("pair_id"): pair_id,
                t("f0_coefficient"): f"{f0:.4f}" if isinstance(f0, (int, float)) else str(f0),
                t("f2_coefficient"): f"{f2:.6f}" if isinstance(f2, (int, float)) else str(f2),
                t("energy"): f"{energy:.4f}" if isinstance(energy, (int, float)) else str(energy),
                t("temperature"): f"{temp:.1f}°C" if isinstance(temp, (int, float)) else str(temp),
                t("pressure"): f"{press:.3f} kPa" if isinstance(press, (int, float)) else str(press),
            })
        
        df_pairs = pd.DataFrame(pairs_data)
        st.dataframe(df_pairs, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # ===== EXPORTAÇÃO =====
    st.subheader(f"📥 {t('export_to_excel')}")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button(f"📊 {t('export_to_excel')}", type="primary", use_container_width=True):
            try:
                export_data = prepare_export_data(t)
                excel_buffer = generate_excel_report(export_data, t)
                st.session_state.excel_buffer = excel_buffer
                st.success(t("export_success"))
            except Exception as e:
                st.error(f"{t('error_calculation')}: {str(e)}")
    
    with col2:
        if "excel_buffer" in st.session_state and st.session_state.excel_buffer:
            st.download_button(
                label=t("download_excel"),
                data=st.session_state.excel_buffer,
                file_name=f"coastdown_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    st.markdown("---")
    
    # ===== RELATÓRIO TEXTUAL =====
    st.subheader("📝 Relatório")
    
    report_text = generate_text_report(t, results)
    st.text_area("Relatório", value=report_text, height=300, disabled=True)
    
    st.download_button(
        label="📄 Baixar Relatório (TXT)",
        data=report_text,
        file_name=f"coastdown_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
        mime="text/plain"
    )


def prepare_export_data(t):
    """Prepara os dados para exportação."""
    
    results = st.session_state.final_results
    vehicle_info = results.get("vehicle_info", {})
    
    export_data = {
        "vehicle_model": vehicle_info.get("model", "N/A"),
        "test_date": str(vehicle_info.get("test_date", "N/A")),
        "total_mass": results.get("total_mass", 0),
        "effective_mass": vehicle_info.get("effective_mass", 0),
        "frontal_area": vehicle_info.get("frontal_area", 0),
        "mean_f0": results.get("mean_f0", 0),
        "mean_f2": results.get("mean_f2", 0),
        "cv_f0": results.get("cv_f0", 0),
        "cv_f2": results.get("cv_f2", 0),
        "mean_energy": results.get("mean_energy", 0),
        "num_pairs": results.get("num_pairs", 0),
        "pairs": normalize_selected_pairs()[0],
    }
    
    return export_data


def generate_excel_report(export_data, t):
    """Gera relatório Excel."""
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Coastdown"
    
    # Estilos
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Título
    ws['A1'] = "RELATÓRIO DE ANÁLISE DE COASTDOWN (ABNT 10312)"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    # Informações do veículo
    ws['A3'] = "INFORMAÇÕES DO VEÍCULO"
    ws['A3'].font = header_font
    
    ws['A4'] = "Modelo:"
    ws['B4'] = export_data.get("vehicle_model", "N/A")
    ws['A5'] = "Data do Teste:"
    ws['B5'] = export_data.get("test_date", "N/A")
    ws['A6'] = "Massa Total (kg):"
    ws['B6'] = export_data.get("total_mass", 0)
    ws['A7'] = "Massa Efetiva (kg):"
    ws['B7'] = export_data.get("effective_mass", 0)
    ws['A8'] = "Área Frontal (m²):"
    ws['B8'] = export_data.get("frontal_area", 0)
    
    # Resultados
    ws['A10'] = "RESULTADOS FINAIS"
    ws['A10'].font = header_font
    
    ws['A11'] = "F0 Médio (N):"
    ws['B11'] = f"{export_data.get('mean_f0', 0):.4f}"
    ws['A12'] = "F2 Médio (N/(km/h)²):"
    ws['B12'] = f"{export_data.get('mean_f2', 0):.6f}"
    ws['A13'] = "CV F0 (%):"
    ws['B13'] = f"{export_data.get('cv_f0', 0):.2f}"
    ws['A14'] = "CV F2 (%):"
    ws['B14'] = f"{export_data.get('cv_f2', 0):.2f}"
    ws['A15'] = "Energia Média (MJ/km):"
    ws['B15'] = f"{export_data.get('mean_energy', 0):.4f}"
    ws['A16'] = "Número de Pares:"
    ws['B16'] = export_data.get("num_pairs", 0)
    
    # Tabela de pares
    ws['A18'] = "PARES SELECIONADOS"
    ws['A18'].font = header_font
    
    headers = ["ID do Par", "F0 (N)", "F2 (N/(km/h)²)", "Energia (MJ/km)", "Temp (°C)", "Pressão (kPa)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=19, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 20
    for pair in export_data.get("pairs", []):
        f0 = _get(pair, "f0_corr", "mean_f0_corrected", "f0corr_mean", "f0_mean")
        f2 = _get(pair, "f2_corr", "mean_f2_corrected", "f2corr_mean", "f2_mean")
        energy = _get(pair, "energy", "mean_energy_corrected")
        temp = pair.get("temp", 25)
        press = pair.get("press", 101.325)
        
        ws.cell(row=row, column=1, value=pair.get("pair_id", "N/A")).border = thin_border
        ws.cell(row=row, column=2, value=f"{f0:.4f}").border = thin_border
        ws.cell(row=row, column=3, value=f"{f2:.6f}").border = thin_border
        ws.cell(row=row, column=4, value=f"{energy:.4f}").border = thin_border
        ws.cell(row=row, column=5, value=f"{temp:.1f}" if isinstance(temp, (int, float)) else str(temp)).border = thin_border
        ws.cell(row=row, column=6, value=f"{press:.3f}" if isinstance(press, (int, float)) else str(press)).border = thin_border
        row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    # Salva em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def generate_text_report(t, results):
    """Gera relatório em texto."""
    
    vehicle_info = results.get("vehicle_info", {})
    
    report = f"""
================================================================================
                    RELATÓRIO DE ANÁLISE DE COASTDOWN
                           (ABNT 10312)
================================================================================

Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

--------------------------------------------------------------------------------
                         INFORMAÇÕES DO VEÍCULO
--------------------------------------------------------------------------------

Modelo:              {vehicle_info.get('model', 'N/A')}
Data do Teste:       {vehicle_info.get('test_date', 'N/A')}
Massa Total:         {results.get('total_mass', 0):.1f} kg
Massa Efetiva:       {vehicle_info.get('effective_mass', 0):.1f} kg
Área Frontal:        {vehicle_info.get('frontal_area', 0):.2f} m²

--------------------------------------------------------------------------------
                          RESULTADOS FINAIS
--------------------------------------------------------------------------------

Coeficiente F0:      {results.get('mean_f0', 0):.4f} N
Coeficiente F2:      {results.get('mean_f2', 0):.6f} N/(km/h)²

CV F0:               {results.get('cv_f0', 0):.2f}%
CV F2:               {results.get('cv_f2', 0):.2f}%

Energia Média:       {results.get('mean_energy', 0):.4f} MJ/km
Número de Pares:     {results.get('num_pairs', 0)}

--------------------------------------------------------------------------------
                          PARES SELECIONADOS
--------------------------------------------------------------------------------

"""
    
    for i, pair in enumerate(normalize_selected_pairs()[0], 1):
        f0 = _get(pair, "f0_corr", "mean_f0_corrected", "f0corr_mean", "f0_mean")
        f2 = _get(pair, "f2_corr", "mean_f2_corrected", "f2corr_mean", "f2_mean")
        energy = _get(pair, "energy", "mean_energy_corrected")
        temp = pair.get("temp", "N/A")
        press = pair.get("press", "N/A")
        
        temp_str = f"{temp:.1f}" if isinstance(temp, (int, float)) else str(temp)
        press_str = f"{press:.3f}" if isinstance(press, (int, float)) else str(press)
        
        report += f"""
Par {i}: {pair.get('pair_id', 'N/A')}
  - F0: {f0:.4f} N
  - F2: {f2:.6f} N/(km/h)²
  - Energia: {energy:.4f} MJ/km
  - Temperatura: {temp_str}°C
  - Pressão: {press_str} kPa
"""
    
    report += """
================================================================================
                              FIM DO RELATÓRIO
================================================================================
"""
    
    return report
