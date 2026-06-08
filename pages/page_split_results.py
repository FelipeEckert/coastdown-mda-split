# coding: utf-8
"""Split final results and Excel export page."""

import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.split_calculations import coefficient_summary


def _result_rows(results: list[dict]) -> list[dict]:
    rows = []
    for idx, result in enumerate(results, start=1):
        high_plus = result.get("high_plus") or result.get("high_record", {})
        low_plus = result.get("low_plus") or result.get("low_record", {})
        high_minus = result.get("high_minus", {})
        low_minus = result.get("low_minus", {})
        rows.append(
            {
                "Index": idx,
                "Use": result.get("selected", True),
                "f'0 (N)": result.get("f0_prime"),
                "f'2 (N/(m/s)^2)": result.get("f2_prime"),
                "High + run": high_plus.get("run_id"),
                "Low + run": low_plus.get("run_id"),
                "High - run": high_minus.get("run_id"),
                "Low - run": low_minus.get("run_id"),
                "High + Delta t": high_plus.get("delta_t_s"),
                "Low + Delta t": low_plus.get("delta_t_s"),
                "High - Delta t": high_minus.get("delta_t_s"),
                "Low - Delta t": low_minus.get("delta_t_s"),
                "Warnings": "; ".join(result.get("warnings", [])),
            }
        )
    return rows


def _write_header(ws, row, values):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def generate_split_excel(results: list[dict], summary: dict, vehicle_info: dict, config: dict) -> bytes:
    """Generate a compact Split Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Split Summary"

    ws["A1"] = "Coastdown MDA Split"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Vehicle model"
    ws["B3"] = vehicle_info.get("model", "N/A")
    ws["A4"] = "Test date"
    ws["B4"] = str(vehicle_info.get("test_date", "N/A"))
    ws["A5"] = "Effective mass (kg)"
    ws["B5"] = vehicle_info.get("effective_mass", "N/A")

    ws["A7"] = "Mean f'0 (N)"
    ws["B7"] = summary.get("mean_f0_prime")
    ws["A8"] = "Mean f'2 (N/(m/s)^2)"
    ws["B8"] = summary.get("mean_f2_prime")
    ws["A9"] = "CV f'0 (%)"
    ws["B9"] = summary.get("cv_f0_prime")
    ws["A10"] = "CV f'2 (%)"
    ws["B10"] = summary.get("cv_f2_prime")
    ws["A11"] = "Number of results"
    ws["B11"] = summary.get("num_results")

    ws["A13"] = "High interval"
    ws["B13"] = f"{config['high']['start']}-{config['high']['end']} km/h; ref {config['high']['reference']} km/h"
    ws["A14"] = "Low interval"
    ws["B14"] = f"{config['low']['start']}-{config['low']['end']} km/h; ref {config['low']['reference']} km/h"

    details = wb.create_sheet("Split Results")
    headers = [
        "Index", "f'0 (N)", "f'2 (N/(m/s)^2)",
        "High + file", "High + run", "High + Delta t (s)", "High + subintervals",
        "Low + file", "Low + run", "Low + Delta t (s)", "Low + subintervals",
        "High - file", "High - run", "High - Delta t (s)", "High - subintervals",
        "Low - file", "Low - run", "Low - Delta t (s)", "Low - subintervals",
        "f'0 + (N)", "f'2 + (N/(m/s)^2)",
        "f'0 - (N)", "f'2 - (N/(m/s)^2)",
        "Warnings",
    ]
    _write_header(details, 1, headers)
    for row_idx, result in enumerate(results, start=2):
        high_plus = result.get("high_plus") or result.get("high_record", {})
        low_plus = result.get("low_plus") or result.get("low_record", {})
        high_minus = result.get("high_minus", {})
        low_minus = result.get("low_minus", {})
        result_plus = result.get("result_plus", {})
        result_minus = result.get("result_minus", {})
        values = [
            row_idx - 1,
            result.get("f0_prime"),
            result.get("f2_prime"),
            high_plus.get("filename"),
            high_plus.get("run_id"),
            high_plus.get("delta_t_s"),
            ", ".join(high_plus.get("subintervals", [])),
            low_plus.get("filename"),
            low_plus.get("run_id"),
            low_plus.get("delta_t_s"),
            ", ".join(low_plus.get("subintervals", [])),
            high_minus.get("filename"),
            high_minus.get("run_id"),
            high_minus.get("delta_t_s"),
            ", ".join(high_minus.get("subintervals", [])),
            low_minus.get("filename"),
            low_minus.get("run_id"),
            low_minus.get("delta_t_s"),
            ", ".join(low_minus.get("subintervals", [])),
            result_plus.get("f0_prime"),
            result_plus.get("f2_prime"),
            result_minus.get("f0_prime"),
            result_minus.get("f2_prime"),
            "; ".join(result.get("warnings", [])),
        ]
        for col_idx, value in enumerate(values, start=1):
            details.cell(row=row_idx, column=col_idx, value=value)

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            sheet.column_dimensions[column_cells[0].column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def render(t):
    """Render Split results."""
    st.header(t("page_split_results"))

    results = st.session_state.get("split_results", [])
    if not results:
        st.info("No Split results have been saved yet.")
        return

    st.subheader("Saved Split results")
    rows = _result_rows(results)
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    selected_results = []
    st.subheader("Select results for final summary")
    for idx, result in enumerate(results):
        label = f"Result {idx + 1}: f'0={result['f0_prime']:.4f} N, f'2={result['f2_prime']:.6f} N/(m/s)^2"
        selected = st.checkbox(label, value=result.get("selected", True), key=f"split_result_selected_{idx}")
        result["selected"] = selected
        if selected:
            selected_results.append(result)

    if st.button("Calculate final Split summary", type="primary", use_container_width=True):
        try:
            summary = coefficient_summary(selected_results)
            summary["vehicle_info"] = st.session_state.vehicle_info
            st.session_state.split_final_results = summary
            st.session_state.excel_buffer = None
            st.success("Final Split summary calculated.")
        except ValueError as exc:
            st.error(str(exc))

    summary = st.session_state.get("split_final_results") or {}
    if not summary:
        return

    st.markdown("---")
    st.subheader("Final summary")
    col1, col2, col3 = st.columns(3)
    col1.metric("Mean f'0", f"{summary['mean_f0_prime']:.4f} N")
    col2.metric("Mean f'2", f"{summary['mean_f2_prime']:.6f} N/(m/s)^2")
    col3.metric("Results", str(summary["num_results"]))

    col4, col5 = st.columns(2)
    col4.metric("CV f'0", f"{summary['cv_f0_prime']:.2f}%")
    col5.metric("CV f'2", f"{summary['cv_f2_prime']:.2f}%")

    export_results = [result for result in results if result.get("selected", True)]
    excel_bytes = generate_split_excel(
        export_results,
        summary,
        st.session_state.vehicle_info,
        st.session_state.split_interval_config,
    )
    st.download_button(
        "Download Split Excel report",
        data=excel_bytes,
        file_name="coastdown_mda_split_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
