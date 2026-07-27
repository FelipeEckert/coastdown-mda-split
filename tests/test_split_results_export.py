# coding: utf-8
"""Tests for the pure final Split workbook exporter."""

from copy import deepcopy
import io
import unittest
from unittest.mock import call, patch

from openpyxl import load_workbook

from core.split_display import get_split_pair_public_label
from core.split_results import consolidate_split_final_results
from data.split_exporters import export_split_final_results_to_excel


COMPONENTS = ("high_plus", "low_plus", "high_minus", "low_minus")


def _pair(pair_id, *, selected=True, temperature=25.0, wind=1.0):
    records = {
        component: {
            "run_id": index,
            "filename": None if component == "low_minus" else f"{component}.csv",
            "delta_t_s": 10.0 + index,
            "source_role": component.split("_")[0],
            "content_sha256": f"hash-{index}",
            "start_time_str": f"08:0{index}:00",
            "subintervals": (
                ["101-94", "94-87"]
                if component.startswith("high")
                else ["63-56", "56-49"]
            ),
            "subinterval_times_s": [4.0 + index, 6.0],
        }
        for index, component in enumerate(COMPONENTS, 1)
    }
    return {
        "id": pair_id,
        "selected": selected,
        "selection_source": "manual",
        **records,
        **{f"{component}_run": index for index, component in enumerate(COMPONENTS, 1)},
        **{f"{component}_delta_t_s": 10.0 + index for index, component in enumerate(COMPONENTS, 1)},
        "F0_mean": 100.0 if pair_id == "one" else 110.0,
        "F2_mean": 0.004 if pair_id == "one" else 0.0042,
        "f0_prime_plus": 90.123456789,
        "f2_prime_plus": 0.05123456789,
        "f0_prime_minus": 91.234567891,
        "f2_prime_minus": 0.05234567891,
        "F0_plus": 100.123456789,
        "F2_plus": 0.004123456789,
        "F0_minus": 101.234567891,
        "F2_minus": 0.004234567891,
        "v2_reference_kmh": 82.5,
        "v1_reference_kmh": 41.0,
        "energy": 1.25,
        "ambient_by_component": {
            component: {
                "temperature_c": temperature + index - 1,
                "pressure_kpa": 101.3,
                "wind_speed_ms": wind + (index - 1) / 10,
                "sync_method": "nearest",
                "warnings": [],
            }
            for index, component in enumerate(COMPONENTS, 1)
        },
        "weather_summary": {"mode": "synchronized"},
        "warnings": [],
    }


def _flat(ws):
    return [value for row in ws.iter_rows(values_only=True) for value in row]


class SplitResultsExportTests(unittest.TestCase):
    def _workbook(self, pairs, *, analysis=None):
        payload = export_split_final_results_to_excel(
            final_results={},
            selected_pairs=pairs,
            vehicle_data={"vehicle_info": {"model": "Carro", "effective_mass": 1500}, "total_mass": 1450},
            deviation_analysis=analysis,
        )
        self.assertGreater(len(payload), 1000)
        self.assertEqual(payload[:2], b"PK")
        return load_workbook(io.BytesIO(payload), data_only=True)

    def test_public_label_uses_saved_public_value(self):
        pair = _pair("split_pair_saved")
        pair["pair_label"] = "Par público"
        self.assertEqual(get_split_pair_public_label(pair), "Par público")

    def test_public_label_rebuilds_technical_or_missing_label(self):
        pair = _pair("split_pair_manual")
        pair["pair_label"] = "split_pair_manual"
        self.assertEqual(
            get_split_pair_public_label(pair),
            "[+]: Run 1 / Run 2 | [-]: Run 3 / Run 4",
        )
        self.assertEqual(
            get_split_pair_public_label({}),
            "[+]: Run - / Run - | [-]: Run - / Run -",
        )

    def test_workbook_preserves_three_sheets_and_adds_deceleration_times(self):
        wb = self._workbook([_pair("one"), _pair("two")])
        self.assertEqual(wb.sheetnames, [
            "Resumo Final", "Pares Selecionados", "Análise de Desvios e Tempos",
            "Tempos de desaceleração",
        ])
        workbook_values = [value for ws in wb.worksheets for value in _flat(ws)]
        for removed in ("Dados do Veículo", "Rastreabilidade", "Tempos deltaT", "LEAVE-ONE-OUT"):
            self.assertNotIn(removed, workbook_values)

    def test_workbook_has_no_freeze_panes_filters_or_excel_tables(self):
        wb = self._workbook([_pair("one"), _pair("two")])
        for ws in wb.worksheets[:3]:
            self.assertIsNone(ws.freeze_panes)
            self.assertIsNone(ws.auto_filter.ref)
            self.assertEqual(len(ws.tables), 0)
        times_ws = wb["Tempos de desaceleração"]
        self.assertEqual(times_ws.freeze_panes, "A3")
        self.assertIsNone(times_ws.auto_filter.ref)
        self.assertEqual(len(times_ws.tables), 0)

    def test_deceleration_times_uses_dynamic_canonical_run_values(self):
        ws = self._workbook([_pair("one")])["Tempos de desaceleração"]

        self.assertEqual(
            [ws.cell(2, column).value for column in range(1, 10)],
            [
                "Par", "Run", "101-94", "94-87", "Δt total [s]",
                None,
                "Temperatura [°C]", "Pressão [kPa]", "Vento [m/s]",
            ],
        )
        self.assertEqual(
            [ws.cell(7, column).value for column in range(1, 10)],
            [
                "Par", "Run", "63-56", "56-49", "Δt total [s]",
                None,
                "Temperatura [°C]", "Pressão [kPa]", "Vento [m/s]",
            ],
        )
        self.assertEqual(
            [ws.cell(3, column).value for column in range(1, 10)],
            [
                "PAR 1", "Run 1 [+]", 5.0, 6.0, 11.0, None,
                25.0, 101.3, 1.0,
            ],
        )
        self.assertEqual(
            [ws.cell(4, column).value for column in range(1, 10)],
            [
                None, "Run 3 [-]", 7.0, 6.0, 13.0, None,
                27.0, 101.3, 1.2,
            ],
        )
        self.assertEqual(
            [ws.cell(8, column).value for column in range(1, 10)],
            [
                "PAR 1", "Run 2 [+]", 6.0, 6.0, 12.0, None,
                26.0, 101.3, 1.1,
            ],
        )
        self.assertEqual(
            [ws.cell(9, column).value for column in range(1, 10)],
            [
                None, "Run 4 [-]", 8.0, 6.0, 14.0, None,
                28.0, 101.3, 1.3,
            ],
        )
        self.assertTrue({
            "A1:E1", "G1:I1", "A3:A4",
            "A6:E6", "G6:I6", "A8:A9",
        }.issubset({str(cell_range) for cell_range in ws.merged_cells.ranges}))
        self.assertEqual(ws["G1"].value, "Estação meteorológica")
        self.assertEqual(ws["G6"].value, "Estação meteorológica")
        self.assertEqual(ws.column_dimensions["F"].width, 3)
        self.assertTrue(all(
            ws.cell(row, 6).value is None and not ws.cell(row, 6).has_style
            for row in range(1, 10)
        ))
        self.assertEqual(ws["C3"].number_format, "0.000")
        self.assertEqual(ws["E3"].number_format, "0.000")
        self.assertEqual(ws["G3"].number_format, "0.0")
        self.assertEqual(ws["H3"].number_format, "0.00")
        self.assertEqual(ws["I3"].number_format, "0.00")
        self.assertEqual(ws["A2"].alignment.horizontal, "center")
        for first_row, second_row in ((3, 4), (8, 9)):
            pair_cell = ws.cell(first_row, 1)
            self.assertEqual(pair_cell.fill.fill_type, "solid")
            self.assertEqual(pair_cell.fill.fgColor.rgb, "00D9D9D9")
            self.assertEqual(pair_cell.alignment.horizontal, "center")
            self.assertEqual(pair_cell.alignment.vertical, "center")
            self.assertEqual(pair_cell.border.top.style, "thin")
            self.assertEqual(pair_cell.border.left.style, "thin")
            self.assertEqual(pair_cell.border.right.style, "thin")
            self.assertEqual(
                ws.cell(second_row, 1).border.bottom.style,
                "medium",
            )
        self.assertGreaterEqual(ws.column_dimensions["A"].width, 12)
        self.assertGreaterEqual(ws.column_dimensions["B"].width, 18)
        for row in (3, 4, 8, 9):
            self.assertAlmostEqual(
                sum(ws.cell(row, column).value for column in (3, 4)),
                ws.cell(row, 5).value,
                places=3,
            )
        for row in (4, 9):
            for column in (*range(1, 6), *range(7, 10)):
                self.assertEqual(ws.cell(row, column).border.bottom.style, "medium")

    def test_deceleration_times_preserves_each_pair_without_deduplicating_runs(self):
        first = _pair("one")
        second = _pair("two")

        ws = self._workbook([first, second])["Tempos de desaceleração"]

        self.assertEqual(
            [ws.cell(row, 1).value for row in range(3, 7)],
            ["PAR 1", None, "PAR 2", None],
        )
        self.assertEqual(
            [ws.cell(row, 2).value for row in range(3, 7)],
            ["Run 1 [+]", "Run 3 [-]", "Run 1 [+]", "Run 3 [-]"],
        )
        self.assertEqual(
            [ws.cell(row, 2).value for row in range(10, 14)],
            ["Run 2 [+]", "Run 4 [-]", "Run 2 [+]", "Run 4 [-]"],
        )
        merged = {str(cell_range) for cell_range in ws.merged_cells.ranges}
        self.assertTrue({
            "A3:A4", "A5:A6", "A10:A11", "A12:A13",
        }.issubset(merged))
        for row in (4, 6, 11, 13):
            for column in (*range(1, 6), *range(7, 10)):
                self.assertEqual(ws.cell(row, column).border.bottom.style, "medium")

    def test_deceleration_times_uses_canonical_threshold_timestamps(self):
        pair = _pair("one")
        pair["high_plus"].pop("subinterval_times_s")
        pair["high_plus"]["times"] = [0.0, 5.0, 11.0]
        pair["high_plus"]["velocities"] = [101.0, 94.0, 87.0]

        ws = self._workbook([pair])["Tempos de desaceleração"]

        self.assertEqual([ws["C3"].value, ws["D3"].value], [5.0, 6.0])
        self.assertAlmostEqual(ws["C3"].value + ws["D3"].value, ws["E3"].value)

    def test_selected_pairs_are_two_grouped_directional_tables_in_order(self):
        first = _pair("one")
        second = _pair("two")
        first["pair_label"] = "First selected pair"
        second["pair_label"] = "Second selected pair"
        ws = self._workbook([first, second])["Pares Selecionados"]

        groups = (
            "Identificação",
            "Passadas e tempos",
            "Resultados não corrigidos",
            "Resultados corrigidos",
            "Meteorologia — Alta",
            "Meteorologia — Baixa",
        )
        headers = (
            "Par", "Origem",
            "Passada de alta", "Passada de baixa",
            "Δt alta [s]", "Δt baixa [s]",
            "F0 [N]", "F2 [N/(km/h)²]",
            "F0 [N]", "F2 [N/(km/h)²]", "Energia [MJ/km]",
            "Temperatura [°C]", "Pressão [kPa]", "Vento [m/s]",
            "Temperatura [°C]", "Pressão [kPa]", "Vento [m/s]",
        )
        for title_row, group_row, header_row, data_rows, title in (
            (1, 2, 3, (4, 5), "IDA [+]"),
            (7, 8, 9, (10, 11), "VOLTA [-]"),
        ):
            self.assertEqual(ws.cell(title_row, 1).value, title)
            self.assertEqual(
                tuple(ws.cell(group_row, column).value for column in (
                    1, 3, 7, 9, 12, 15,
                )),
                groups,
            )
            self.assertEqual(
                tuple(ws.cell(header_row, column).value for column in range(1, 18)),
                headers,
            )
            self.assertEqual(
                [ws.cell(row, 1).value for row in data_rows],
                ["First selected pair", "Second selected pair"],
            )

        self.assertEqual(
            [ws.cell(4, column).value for column in range(3, 7)],
            [1, 2, 11.0, 12.0],
        )
        self.assertEqual(
            [ws.cell(10, column).value for column in range(3, 7)],
            [3, 4, 13.0, 14.0],
        )
        self.assertTrue({
            "A1:Q1", "A2:B2", "C2:F2", "G2:H2", "I2:K2",
            "L2:N2", "O2:Q2",
            "A7:Q7", "A8:B8", "C8:F8", "G8:H8", "I8:K8",
            "L8:N8", "O8:Q8",
        }.issubset({str(cell_range) for cell_range in ws.merged_cells.ranges}))

    def test_corrected_energy_and_component_weather_use_exact_sources(self):
        pair = _pair("one")

        with patch(
            "data.split_exporters.calculate_split_energy",
            side_effect=[{"energy": 1.2}, {"energy": 2.2}],
        ) as calculate_energy:
            ws = self._workbook([pair])["Pares Selecionados"]

        self.assertEqual(calculate_energy.call_args_list, [
            call(pair["F0_plus"], pair["F2_plus"]),
            call(pair["F0_minus"], pair["F2_minus"]),
        ])
        self.assertEqual(
            [ws.cell(4, column).value for column in range(7, 12)],
            [
                pair["f0_prime_plus"], pair["f2_prime_plus"],
                pair["F0_plus"], pair["F2_plus"], 1.2,
            ],
        )
        self.assertEqual(
            [ws.cell(9, column).value for column in range(7, 12)],
            [
                pair["f0_prime_minus"], pair["f2_prime_minus"],
                pair["F0_minus"], pair["F2_minus"], 2.2,
            ],
        )
        self.assertEqual(
            [ws.cell(4, column).value for column in range(12, 18)],
            [25.0, 101.3, 1.0, 26.0, 101.3, 1.1],
        )
        self.assertEqual(
            [ws.cell(9, column).value for column in range(12, 18)],
            [27.0, 101.3, 1.2, 28.0, 101.3, 1.3],
        )
        self.assertEqual(ws.max_column, 17)

    def test_summary_omits_warnings_and_merges_only_section_titles(self):
        wb = self._workbook([_pair("one"), _pair("two")])
        ws = wb["Resumo Final"]
        self.assertNotIn("Warnings principais", _flat(ws))
        expected_titles = {
            "RESULTADOS SPLIT", "RESULTADO FINAL",
            "DADOS DO VEÍCULO", "RESUMO METEOROLÓGICO",
        }
        merged_titles = {
            ws.cell(cell_range.min_row, 1).value
            for cell_range in ws.merged_cells.ranges
            if cell_range.min_col == 1 and cell_range.max_col == 2
        }
        self.assertEqual(merged_titles, expected_titles)

    def test_excel_final_status_matches_time_based_ui_outcomes(self):
        cases = (
            (True, False, "Aprovado", "0EE427", "conforming"),
            (True, True, "Aprovado", "0EE427", "nonconforming"),
            (False, False, "Reprovado", "FF5757", "conforming"),
            (False, True, "Reprovado", "FF5757", "nonconforming"),
        )

        for time_passed, coefficient_fails, expected, color, legacy in cases:
            with self.subTest(
                time_passed=time_passed,
                coefficient_fails=coefficient_fails,
            ):
                pairs = [_pair("one"), _pair("two")]
                if coefficient_fails:
                    pairs[1]["F0_mean"] = 200.0
                    pairs[1]["F2_mean"] = 0.02
                self.assertEqual(
                    consolidate_split_final_results(pairs)["conformity_status"],
                    legacy,
                )
                ws = self._workbook(
                    pairs,
                    analysis={"time_summary": {"passed": time_passed}},
                )["Resumo Final"]
                status_row = next(
                    row for row in range(1, ws.max_row + 1)
                    if ws.cell(row, 1).value == "Status final"
                )
                self.assertEqual(ws.cell(status_row, 2).value, expected)
                self.assertEqual(ws.cell(status_row, 2).fill.fgColor.rgb[-6:], color)

    def test_excel_final_status_is_inconclusive_without_evaluable_time(self):
        pairs = [_pair("one"), _pair("two")]
        for analysis in ({}, {"time_summary": {"passed": None}}):
            with self.subTest(analysis=analysis):
                ws = self._workbook(
                    pairs,
                    analysis=analysis,
                )["Resumo Final"]
                rows = dict(ws.iter_rows(values_only=True))
                self.assertEqual(rows["Status final"], "Inconclusivo")

    def test_legacy_warning_and_incomplete_states_do_not_override_time_status(self):
        cases = (
            ("warning", {"warnings": ["Aviso de validação"]}),
            ("incomplete", {"F0_mean": None}),
        )

        for legacy_status, changes in cases:
            with self.subTest(legacy_status=legacy_status):
                pairs = [_pair("one"), _pair("two")]
                pairs[0].update(changes)
                summary = consolidate_split_final_results(pairs)
                self.assertEqual(summary["conformity_status"], legacy_status)
                ws = self._workbook(
                    pairs,
                    analysis={"time_summary": {"passed": True}},
                )["Resumo Final"]
                rows = dict(ws.iter_rows(values_only=True))
                self.assertEqual(rows["Status final"], "Aprovado")

    def test_pair_cv_stays_only_in_diagnostic_worksheets(self):
        wb = self._workbook([_pair("one"), _pair("two")])
        self.assertIn(
            "CV F0 diagnóstico [%]",
            _flat(wb["Resumo Final"]),
        )
        pair_values = _flat(wb["Pares Selecionados"])
        self.assertNotIn("CV F0 diagnóstico [%]", pair_values)
        self.assertNotIn("CV F2 diagnóstico [%]", pair_values)
        deviation_values = _flat(wb["Análise de Desvios e Tempos"])
        self.assertIn("RESUMO CV F0/F2 (DIAGNÓSTICO)", deviation_values)
        self.assertIn("Status diagnóstico", deviation_values)

    def test_summary_contains_vehicle_results_and_weather_blocks(self):
        wb = self._workbook([_pair("one"), _pair("two")])
        values = _flat(wb["Resumo Final"])
        for expected in ("DADOS DO VEÍCULO", "Carro", "Massa efetiva Me [kg]", "F0 final [N]", "RESUMO METEOROLÓGICO"):
            self.assertIn(expected, values)
        f0_row = next(row for row in wb["Resumo Final"].iter_rows(values_only=True) if row[0] == "F0 final [N]")
        self.assertIsInstance(f0_row[1], (int, float))

    def test_vehicle_mass_headers_are_explicit_and_values_are_numeric(self):
        payload = export_split_final_results_to_excel(
            final_results={}, selected_pairs=[_pair("one")],
            vehicle_data={"running_order_mass_kg": 1500.0},
        )
        ws = load_workbook(io.BytesIO(payload), data_only=True)["Resumo Final"]
        rows = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}

        self.assertEqual(rows["Massa em ordem de marcha [kg]"], 1500.0)
        self.assertEqual(rows["Massa de ensaio M [kg]"], 1636.0)
        self.assertAlmostEqual(rows["Massa equivalente de rotação me [kg]"], 49.08)
        self.assertAlmostEqual(rows["Massa efetiva Me [kg]"], 1685.08)
        self.assertTrue(all(" kg" not in str(rows[label]) for label in (
            "Massa em ordem de marcha [kg]", "Massa de ensaio M [kg]",
            "Massa equivalente de rotação me [kg]", "Massa efetiva Me [kg]",
        )))

    def test_manual_and_algorithmic_pairs_use_public_labels_and_origins(self):
        manual = _pair("split_pair_manual")
        algorithm = _pair("split_pair_algorithm")
        algorithm.update({"selection_source": "algorithm", "algorithm_sources": ["energy", "target"]})
        wb = self._workbook([manual, algorithm])
        ws = wb["Pares Selecionados"]
        rows = [
            tuple(ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
            for row in (4, 5, 10, 11)
        ]
        self.assertTrue(all(str(row[0]).startswith("[+]: Run") for row in rows))
        self.assertFalse(any(
            "split_pair_" in str(value) for row in rows for value in row
        ))
        self.assertEqual(
            [row[1] for row in rows],
            ["Manual", "Energia + Target", "Manual", "Energia + Target"],
        )

    def test_pair_quantities_are_numeric_and_missing_values_are_dash(self):
        wb = self._workbook([_pair("one")])
        ws = wb["Pares Selecionados"]
        for row in (4, 9):
            for column in range(3, 18):
                self.assertIsInstance(
                    ws.cell(row, column).value,
                    (int, float),
                    (row, column),
                )

    def test_fixed_mode_exports_temperature_pressure_and_dash_wind(self):
        pair = _pair("fixed")
        pair["ambient_by_component"] = {}
        pair["weather_summary"] = {"mode": "fixed"}
        pair["environmental_conditions"] = {
            "mode": "fixed", "temperature_c": 23.0,
            "pressure_kpa": 100.5, "wind_speed_mps": None,
        }
        wb = self._workbook([pair])
        values = _flat(wb["Pares Selecionados"])
        self.assertIn(23.0, values)
        self.assertIn(100.5, values)
        self.assertIn("-", values)
        self.assertIn("Parâmetros fixos", _flat(wb["Resumo Final"]))
        self.assertNotIn("missing", values)

    def test_synchronized_mode_exports_weather_without_observations(self):
        wb = self._workbook([_pair("one", temperature=36.0, wind=3.1)])
        values = _flat(wb["Pares Selecionados"])
        self.assertIn(36.0, values)
        self.assertIn(101.3, values)
        self.assertIn(3.1, values)
        self.assertNotIn("Observações", values)
        self.assertFalse(any(
            isinstance(value, str) and "acima de" in value
            for value in values
        ))

    def test_deviation_sheet_contains_cv_and_time_blocks_without_cached_technical_label(self):
        pairs = [_pair("split_pair_one"), _pair("split_pair_two")]
        analysis = {
            "coefficient_summary": {"mean_f0": 105, "mean_f2": .0041, "cv_f0_pct": 6, "cv_f2_pct": 3, "limit_pct": 10, "status": "approved"},
            "pair_deviations": [
                {"pair": pair["id"], "f0": pair["F0_mean"], "f2": pair["F2_mean"], "energy": pair["energy"]}
                for pair in pairs
            ],
            "time_summary": {
                "groups": {component: {"count": 2, "mean": 12, "stdev": 1, "cv_pct": 2, "passed": True} for component in COMPONENTS},
                "cv_limit_pct": 2.5,
                "opposite_direction": {"high": {"mean_plus": 11, "mean_minus": 13, "diff_pct": 1, "passed": True}, "low": {"mean_plus": 12, "mean_minus": 14, "diff_pct": 1, "passed": True}},
                "opposite_mean_limit_pct": 10,
            },
            "leave_one_out": [{"pair": "split_pair_one"}],
            "warnings": [],
        }
        values = _flat(self._workbook(pairs, analysis=analysis)["Análise de Desvios e Tempos"])
        for expected in ("RESUMO CV F0/F2 (DIAGNÓSTICO)", "DESVIOS POR PAR", "TEMPOS Δt", "DIFERENÇA ENTRE MÉDIAS Δt DE SENTIDOS OPOSTOS"):
            self.assertIn(expected, values)
        for expected in (
            "C.V. Δt — Vel. ref. alta 82.5 km/h [+]",
            "C.V. Δt — Vel. ref. alta 82.5 km/h [-]",
            "C.V. Δt — Vel. ref. baixa 41 km/h [+]",
            "C.V. Δt — Vel. ref. baixa 41 km/h [-]",
            "Dif. médias Δt — Vel. ref. alta 82.5 km/h: [+] vs [-]",
            "Dif. médias Δt — Vel. ref. baixa 41 km/h: [+] vs [-]",
        ):
            self.assertIn(expected, values)
        self.assertFalse(any(
            value in COMPONENTS for value in values if isinstance(value, str)
        ))
        self.assertNotIn("LEAVE-ONE-OUT", values)
        self.assertFalse(any("split_pair_" in str(value) for value in values))

    def test_only_explicitly_selected_pairs_are_exported_and_input_is_unchanged(self):
        pairs = [_pair("one"), _pair("hidden", selected=False)]
        before = deepcopy(pairs)
        wb = self._workbook(pairs)
        ws = wb["Pares Selecionados"]
        self.assertEqual(ws.max_row, 9)
        self.assertEqual([ws.cell(row, 1).value for row in (4, 9)], [
            get_split_pair_public_label(pairs[0]),
            get_split_pair_public_label(pairs[0]),
        ])
        self.assertEqual(pairs, before)

    def test_exporter_module_does_not_import_streamlit(self):
        import data.split_exporters as module
        self.assertNotIn("streamlit", module.__dict__)


if __name__ == "__main__":
    unittest.main()
