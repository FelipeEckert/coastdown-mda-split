# coding: utf-8
"""
Funções de exportação de dados para análise de coastdown.

Este módulo contém todas as funções de exportação para Excel e templates.

IMPORTANTE: Não altere os nomes das funções ou variáveis para manter compatibilidade.
"""

import os
import statistics
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.calculations import calcular_energia


# Mapeamento de células do template
celulas_template = {
    # Massas e data
    "massa_motorista": "C26",
    "massa_equipamento": "C27",
    "massa_veiculo": "C28",
    "massa_rotacional": "C29",
    "massa_efetiva": "C30",
    "data_teste": "C34",

    # Coeficientes sem correção (5 pares)
    "f0_ida_1": "F107", "f0_volta_1": "F108",
    "f2_ida_1": "H107", "f2_volta_1": "H108",
    "media_f0_1": "J107", "media_f2_1": "L107",

    "f0_ida_2": "F110", "f0_volta_2": "F111",
    "f2_ida_2": "H110", "f2_volta_2": "H111",
    "media_f0_2": "J110", "media_f2_2": "L110",

    "f0_ida_3": "F113", "f0_volta_3": "F114",
    "f2_ida_3": "H113", "f2_volta_3": "H114",
    "media_f0_3": "J113", "media_f2_3": "L113",

    "f0_ida_4": "F116", "f0_volta_4": "F117",
    "f2_ida_4": "H116", "f2_volta_4": "H117",
    "media_f0_4": "J116", "media_f2_4": "L116",

    "f0_ida_5": "F119", "f0_volta_5": "F120",
    "f2_ida_5": "H119", "f2_volta_5": "H120",
    "media_f0_5": "J119", "media_f2_5": "L119",

    # CVs
    "cv_f0": "H123",
    "cv_f2": "O123",

    # Coeficientes corrigidos (5 pares)
    "media_f0_corr_1": "F137", "media_f2_corr_1": "J137", "energia_1": "N137",
    "media_f0_corr_2": "F141", "media_f2_corr_2": "J141", "energia_2": "N141",
    "media_f0_corr_3": "F145", "media_f2_corr_3": "J145", "energia_3": "N145",
    "media_f0_corr_4": "F149", "media_f2_corr_4": "J149", "energia_4": "N149",
    "media_f0_corr_5": "F153", "media_f2_corr_5": "J153", "energia_5": "N153"
}


def preencher_template_com_dados(template_path, output_path, vehicle_info, final_results, selected_pairs_details, all_run_data):
    """
    Preenche um template Excel com os dados de análise de coastdown.
    
    Args:
        template_path: Caminho do template Excel
        output_path: Caminho para salvar o arquivo preenchido
        vehicle_info: Informações do veículo
        final_results: Resultados finais calculados
        selected_pairs_details: Detalhes dos pares selecionados
        all_run_data: Dados de todas as passagens
    """
    wb = load_workbook(template_path, data_only=True)
    ws = wb.active

    def preencher(celula, valor, formato=None):
        try:
            cell = ws[celula]
            cell.value = valor if valor is not None else "N/A"
            if formato:
                cell.number_format = formato
        except Exception as e:
            print(f"[DEBUG] Erro ao preencher célula {celula}: {e}")

    def formatar_cv(valor):
        return f"{valor:.2f}%" if valor is not None else "N/A"
        
    celulas_tempos = {}

    for par in range(1, 6):
        linha_ida = 75 + (par - 1) * 2 + 1
        linha_volta = linha_ida + 1
        velocidades = [95, 90, 85, 80, 75, 70, 65, 60, 55, 50, 45, 40, 35, 30]
        colunas = list("CDEFGHIJKLMNOP")

        for i, vel in enumerate(velocidades):
            celulas_tempos[f"par{par}_ida_vel{vel}"] = f"{colunas[i]}{linha_ida}"
            celulas_tempos[f"par{par}_volta_vel{vel}"] = f"{colunas[i]}{linha_volta}"

        celulas_tempos[f"par{par}_ida_temp"] = f"R{linha_ida}"
        celulas_tempos[f"par{par}_ida_press"] = f"S{linha_ida}"
        celulas_tempos[f"par{par}_ida_wind"] = f"T{linha_ida}"

        celulas_tempos[f"par{par}_volta_temp"] = f"R{linha_volta}"
        celulas_tempos[f"par{par}_volta_press"] = f"S{linha_volta}"
        celulas_tempos[f"par{par}_volta_wind"] = f"T{linha_volta}"
    

    # Massas
    preencher("C26", vehicle_info.get("mass_driver"), "0.0")
    preencher("C27", vehicle_info.get("mass_equip"), "0.0")
    preencher("C28", vehicle_info.get("mass_vehicle"), "0.0")
    preencher("C29", vehicle_info.get("mass_rot"), "0.0")
    preencher("C30", vehicle_info.get("mass_total"), "0.0")

    # Data do teste
    data_teste = vehicle_info.get("date")
    preencher("C34", data_teste if data_teste else "N/A")

    # CVs NÃO CORRIGIDOS com símbolo de porcentagem
    preencher("H123", formatar_cv(final_results.get("cv_f0_raw")))
    preencher("O123", formatar_cv(final_results.get("cv_f2_raw")))


    # Coeficientes dos pares (até 5)
    for i in range(1, 6):
        pair_key = f"{i}"
        if i <= len(selected_pairs_details):
            pair = selected_pairs_details[i - 1]

            # Sem correção
            preencher(celulas_template[f"f0_ida_{pair_key}"], pair.get("f0_ida"), "0.0000")
            preencher(celulas_template[f"f0_volta_{pair_key}"], pair.get("f0_volta"), "0.0000")
            preencher(celulas_template[f"f2_ida_{pair_key}"], pair.get("f2_ida"), "0.000000")
            preencher(celulas_template[f"f2_volta_{pair_key}"], pair.get("f2_volta"), "0.000000")
            preencher(celulas_template[f"media_f0_{pair_key}"], pair.get("f0_mean"), "0.0000")
            preencher(celulas_template[f"media_f2_{pair_key}"], pair.get("f2_mean"), "0.000000")

            # Com correção
            preencher(celulas_template[f"media_f0_corr_{pair_key}"], pair.get("f0_corr"), "0.0000")
            preencher(celulas_template[f"media_f2_corr_{pair_key}"], pair.get("f2_corr"), "0.000000")
            preencher(celulas_template[f"energia_{pair_key}"], pair.get("energy"), "0.0000")
        else:
            for chave in [f"f0_ida_{pair_key}", f"f0_volta_{pair_key}", f"f2_ida_{pair_key}", f"f2_volta_{pair_key}",
                          f"media_f0_{pair_key}", f"media_f2_{pair_key}",
                          f"media_f0_corr_{pair_key}", f"media_f2_corr_{pair_key}", f"energia_{pair_key}"]:
                preencher(celulas_template[chave], "N/A")


    # Médias finais dos pares selecionados
    preencher("E159", final_results.get("mean_f0"), "0.0000")
    preencher("K159", final_results.get("mean_f2"), "0.0000")
    preencher("O159", final_results.get("energy"), "0.0000")

    # Preencher tabela de tempos de desaceleração
    preencher_tabela_tempos(ws, selected_pairs_details, all_run_data, celulas_tempos)

    # Salvar o arquivo
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"[DEBUG] Template preenchido salvo em: {output_path}")
    except Exception as e:
        print(f"[DEBUG] Erro ao salvar o arquivo: {e}")


def preencher_tabela_tempos(ws, pares_finais_selecionados, all_run_data, celulas_tempos):
    """
    Preenche a tabela de tempos de desaceleração no template.
    
    Args:
        ws: Worksheet do openpyxl
        pares_finais_selecionados: Lista de pares selecionados
        all_run_data: Dados de todas as passagens
        celulas_tempos: Mapeamento de células para tempos
    """
    for idx, pair in enumerate(pares_finais_selecionados, start=1):
        for sentido, run_key in [("ida", "run1"), ("volta", "run2")]:
            run_id = pair.get(run_key)
            run_data = all_run_data.get(run_id, {})
            times = run_data.get("times", [])
            velocities = run_data.get("velocities", [])
            temp = pair.get("temp", "N/A")
            press = pair.get("press", "N/A")


            if not times or not velocities or len(times) != len(velocities):
                continue

            # Preenche células com tempo acumulado
            for i, time_val in enumerate(times):
                if i >= len(velocities):
                    break
                vel = int(velocities[i])
                cell_key = f"par{idx}_{sentido}_vel{vel}"
                cell_ref = celulas_tempos.get(cell_key)
                if cell_ref:
                    ws[cell_ref].value = round(time_val, 2)
                    ws[cell_ref].number_format = "0.00"

            # Temperatura e pressão
            ws[celulas_tempos.get(f"par{idx}_{sentido}_temp", "")].value = temp
            ws[celulas_tempos.get(f"par{idx}_{sentido}_press", "")].value = press
            
            # === NOVO: vento por sentido ===
            wind = None
            if sentido == "ida":
                wind = pair.get("wind_ida_ms", None)
            elif sentido == "volta":
                wind = pair.get("wind_volta_ms", None)

            # fallback: se não houver por-run, usa média do par (se existir)
            if wind is None:
                wind = pair.get("wind_avg_ms", None)

            cell_wind_ref = celulas_tempos.get(f"par{idx}_{sentido}_wind", "")
            if cell_wind_ref:
                try:
                    if wind is not None and not pd.isna(wind):
                        wind_val = float(wind)
                        ws[cell_wind_ref].value = round(wind_val, 2)
                        ws[cell_wind_ref].number_format = "0.00"
                    else:
                        ws[cell_wind_ref].value = "N/A"
                except Exception:
                    ws[cell_wind_ref].value = "N/A"
                
            print(f"[DEBUG] Par {idx} - {sentido}: temp={temp}, press={press}, pair_keys={list(pair.keys())}")


def gerar_excel(output_path, filename, data_info, vehicle_info, final_results, selected_pairs_details, all_run_data):
    """
    Gera um arquivo Excel com três abas contendo os resultados da análise de coastdown,
    adaptado do código original para usar as variáveis do sistema atual.
    
    Args:
        output_path: Caminho onde o arquivo será salvo
        filename: Nome do arquivo
        data_info: Informações sobre os dados de teste
        vehicle_info: Informações sobre o veículo
        final_results: Resultados finais calculados
        selected_pairs_details: Detalhes dos pares selecionados para o cálculo final
        all_run_data: Dados de todas as passagens
    
    Returns:
        Caminho completo do arquivo salvo
    """
    # Cria o workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove a planilha padrão

    # Estilos
    fonte_padrao = Font(name='Arial', size=10)
    fonte_negrito = Font(name='Arial', size=10, bold=True)
    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    borda_fina = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    fundo_cinza = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Borda com topo duplo (para separar as linhas de MÉDIAS)
    borda_fina_topo_duplo = Border(
        left=borda_fina.left,
        right=borda_fina.right,
        bottom=borda_fina.bottom,
        top=Side(border_style='double', color="000000")
    )

    def safe_mean(values):
        nums = [v for v in values if isinstance(v, (int, float))]
        return sum(nums) / len(nums) if nums else None

    # ===== ABA 1: RESULTADOS COASTDOWN =====
    ws1 = wb.create_sheet(title="Resultados Coastdown")

    # Cabeçalho geral
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    cell = ws1.cell(row=1, column=1)
    cell.value = (
        f"Resultados Coastdown - {data_info.get('filename', 'N/A')} - "
        f"Veículo: {vehicle_info.get('vehicle_id', 'N/A')}"
    )
    cell.font = Font(name='Arial', size=14, bold=True)
    cell.alignment = alinhamento_centro

    # ---------- helpers locais ----------
    def write_val(ws, row, col, value, num_fmt=None, bold=False, fill=None):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = fonte_negrito if bold else fonte_padrao
        c.alignment = alinhamento_centro
        c.border = borda_fina
        if fill is not None:
            c.fill = fill
        if num_fmt:
            c.number_format = num_fmt
        return c

    def build_run_label(run):
        """Ex.: 1 ↑ [+]  /  2 ↓ [-]"""
        if run in (None, "", "N/A"):
            return str(run)
        run_str = str(run)
        try:
            run_int = int(run)
        except (TypeError, ValueError):
            return run_str

        heading = None
        if all_run_data and run_int in all_run_data:
            heading = all_run_data[run_int].get("heading")

        if heading == "+":
            return f"{run_str} ↑ [+]"
        elif heading == "-":
            return f"{run_str} ↓ [-]"
        else:
            return run_str

    # Cabeçalhos da tabela principal
    headers = [
        "",                      # 1ª coluna sem "Par"
        "Runs",
        "Temperatura (°C)",
        "Pressão (kPa)",
        "Vento (m/s)",
        "f'0",
        "f'2",
        "CV f'0 (%)",
        "CV f'2 (%)",
        "F0 Corr.",
        "F2 Corr.",
        "Energia (MJ/km)",
    ]
    for col, header in enumerate(headers, start=1):
        c = ws1.cell(row=3, column=col)
        c.value = header
        c.font = fonte_negrito
        c.fill = fundo_cinza
        c.alignment = alinhamento_centro
        c.border = borda_fina

    # ---------- Dados dos pares ----------
    current_row = 4

    for idx, pair_info in enumerate(selected_pairs_details, start=1):
        row_ida   = current_row
        row_volta = current_row + 1
        row_media = current_row + 2

        # PAR X (mescla ida/volta)
        ws1.merge_cells(start_row=row_ida, start_column=1, end_row=row_volta, end_column=1)
        write_val(ws1, row_ida, 1, f"PAR {idx}", bold=True, fill=fundo_cinza)
        write_val(ws1, row_media, 1, "Médias", bold=True, fill=fundo_cinza)

        # Runs
        run_ida   = pair_info.get("run1", pair_info.get("run_ida"))
        run_volta = pair_info.get("run2", pair_info.get("run_volta"))

        write_val(ws1, row_ida,   2, build_run_label(run_ida))
        write_val(ws1, row_volta, 2, build_run_label(run_volta))
        write_val(ws1, row_media, 2, "")

        # Temperatura / Pressão (ida/volta + médias)
        temp_ida   = pair_info.get("temp_ida_used", pair_info.get("temp"))
        temp_volta = pair_info.get("temp_volta_used", pair_info.get("temp"))
        press_ida  = pair_info.get("press_ida_used", pair_info.get("press"))
        press_volta= pair_info.get("press_volta_used", pair_info.get("press"))

        temps = [t for t in (temp_ida, temp_volta) if isinstance(t, (int, float))]
        presses = [p for p in (press_ida, press_volta) if isinstance(p, (int, float))]

        temp_mean = statistics.mean(temps) if temps else None
        press_mean = statistics.mean(presses) if presses else None

        write_val(ws1, row_ida,   3, temp_ida   if isinstance(temp_ida, (int, float)) else "")
        write_val(ws1, row_volta, 3, temp_volta if isinstance(temp_volta, (int, float)) else "")
        write_val(ws1, row_media, 3,
                  temp_mean if isinstance(temp_mean, (int, float)) else "",
                  num_fmt="0.00")

        write_val(ws1, row_ida,   4, press_ida   if isinstance(press_ida, (int, float)) else "")
        write_val(ws1, row_volta, 4, press_volta if isinstance(press_volta, (int, float)) else "")
        write_val(ws1, row_media, 4,
                  press_mean if isinstance(press_mean, (int, float)) else "",
                  num_fmt="0.00")

        # Vento (ida/volta + média)
        w1 = pair_info.get("wind_ida_ms")
        w2 = pair_info.get("wind_volta_ms")
        winds = [w for w in (w1, w2) if isinstance(w, (int, float))]
        wind_mean = statistics.mean(winds) if winds else None

        write_val(ws1, row_ida,   5, w1 if isinstance(w1, (int, float)) else "N/A")
        write_val(ws1, row_volta, 5, w2 if isinstance(w2, (int, float)) else "N/A")
        write_val(ws1, row_media, 5,
                  wind_mean if isinstance(wind_mean, (int, float)) else "N/A",
                  num_fmt="0.00")

        # --- Coeficientes individuais (brutos e corrigidos) + médias / CV / energia ---
        # Brutos individuais (por passada)
        f0_ida_raw   = pair_info.get("f0_ida_raw")
        f2_ida_raw   = pair_info.get("f2_ida_raw")
        f0_volta_raw = pair_info.get("f0_volta_raw")
        f2_volta_raw = pair_info.get("f2_volta_raw")

        # Corrigidos individuais (por passada) – podem ser None se não houve correção
        f0_ida_corr   = pair_info.get("f0_ida_corr")
        f2_ida_corr   = pair_info.get("f2_ida_corr")
        f0_volta_corr = pair_info.get("f0_volta_corr")
        f2_volta_corr = pair_info.get("f2_volta_corr")

        # Médias (brutas + corrigidas), CV e energia
        f0_mean      = pair_info.get("f0_mean")   # média BRUTA
        f2_mean      = pair_info.get("f2_mean")
        f0_corr_mean = pair_info.get("f0_corr")   # média CORRIGIDA
        f2_corr_mean = pair_info.get("f2_corr")
        cv_f0        = pair_info.get("cv_f0")
        cv_f2        = pair_info.get("cv_f2")
        energy       = pair_info.get("energy")

        # Energia individual por passada (se tiver coeficientes corrigidos)
        energy_ida = None
        energy_volta = None
        try:
            if isinstance(f0_ida_corr, (int, float)) and isinstance(f2_ida_corr, (int, float)):
                energy_ida = calcular_energia(f0_ida_corr, f2_ida_corr)
            if isinstance(f0_volta_corr, (int, float)) and isinstance(f2_volta_corr, (int, float)):
                energy_volta = calcular_energia(f0_volta_corr, f2_volta_corr)
        except Exception:
            # se por algum motivo der erro, deixa como None
            energy_ida = None
            energy_volta = None

        # --- Linha IDA ---
        write_val(ws1, row_ida, 6,
                  f0_ida_raw if isinstance(f0_ida_raw, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_ida, 7,
                  f2_ida_raw if isinstance(f2_ida_raw, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_ida, 8, "")   # CV por passada não faz sentido
        write_val(ws1, row_ida, 9, "")

        write_val(ws1, row_ida, 10,
                  f0_ida_corr if isinstance(f0_ida_corr, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_ida, 11,
                  f2_ida_corr if isinstance(f2_ida_corr, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_ida, 12,
                  energy_ida if isinstance(energy_ida, (int, float)) else "",
                  num_fmt="0.0000")

        # --- Linha VOLTA ---
        write_val(ws1, row_volta, 6,
                  f0_volta_raw if isinstance(f0_volta_raw, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_volta, 7,
                  f2_volta_raw if isinstance(f2_volta_raw, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_volta, 8, "")
        write_val(ws1, row_volta, 9, "")

        write_val(ws1, row_volta, 10,
                  f0_volta_corr if isinstance(f0_volta_corr, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_volta, 11,
                  f2_volta_corr if isinstance(f2_volta_corr, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_volta, 12,
                  energy_volta if isinstance(energy_volta, (int, float)) else "",
                  num_fmt="0.0000")

        # --- Linha MÉDIAS ---
        write_val(ws1, row_media, 6,
                  f0_mean if isinstance(f0_mean, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_media, 7,
                  f2_mean if isinstance(f2_mean, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_media, 8,
                  cv_f0 if isinstance(cv_f0, (int, float)) else "",
                  num_fmt="0.00")
        write_val(ws1, row_media, 9,
                  cv_f2 if isinstance(cv_f2, (int, float)) else "",
                  num_fmt="0.00")
        write_val(ws1, row_media, 10,
                  f0_corr_mean if isinstance(f0_corr_mean, (int, float)) else "",
                  num_fmt="0.0000")
        write_val(ws1, row_media, 11,
                  f2_corr_mean if isinstance(f2_corr_mean, (int, float)) else "",
                  num_fmt="0.000000")
        write_val(ws1, row_media, 12,
                  energy if isinstance(energy, (int, float)) else "",
                  num_fmt="0.0000")

        # Borda dupla separando volta x médias
        for col in range(1, 13):
            c = ws1.cell(row=row_media, column=col)
            c.border = Border(
                left=c.border.left,
                right=c.border.right,
                top=Side(border_style="double"),
                bottom=c.border.bottom,
            )

        # bloco ocupa 3 linhas (ida, volta, médias) + 1 linha em branco entre pares
        current_row += 4

    # ===== RESULTADOS FINAIS =====
    linha_rf = current_row  # primeira linha livre depois dos pares

    ws1.merge_cells(start_row=linha_rf, start_column=1, end_row=linha_rf, end_column=2)
    titulo_rf = ws1.cell(row=linha_rf, column=1)
    titulo_rf.value = "RESULTADOS FINAIS"
    titulo_rf.font = Font(name='Arial', size=12, bold=True)
    titulo_rf.alignment = alinhamento_centro
    titulo_rf.fill = fundo_cinza
    for col in range(1, 3):
        ws1.cell(row=linha_rf, column=col).border = borda_fina

    resultados_finais = [
        ("Média F0 Corr. (N)",              final_results.get("mean_f0", 0)),
        ("Média F2 Corr. (N/(km/h)²)",      final_results.get("mean_f2", 0)),
        ("Média Energia (MJ/km)",           final_results.get("energy", 0)),
        ("Média CV F0 (%)",                 final_results.get("cv_f0", 0)),
        ("Média CV F2 (%)",                 final_results.get("cv_f2", 0)),
    ]

    for i, (descricao, valor) in enumerate(resultados_finais):
        row = linha_rf + 1 + i
        ws1.cell(row=row, column=1, value=descricao).font = fonte_negrito
        cell_val = ws1.cell(row=row, column=2, value=valor)
        cell_val.font = fonte_padrao
        cell_val.alignment = alinhamento_centro

        # Formatação numérica
        if i == 0:          # Média F0
            cell_val.number_format = "0.0000"
        elif i == 1:        # Média F2
            cell_val.number_format = "0.000000"
        elif i == 2:        # Energia
            cell_val.number_format = "0.0000"
        elif i in (3, 4):   # CVs
            cell_val.number_format = "0.000"

        for col in range(1, 3):
            ws1.cell(row=row, column=col).border = borda_fina

    # ===== INFORMAÇÕES DO ENSAIO =====
    linha_ie = linha_rf + 1 + len(resultados_finais) + 2

    ws1.merge_cells(start_row=linha_ie, start_column=1, end_row=linha_ie, end_column=2)
    titulo_ie = ws1.cell(row=linha_ie, column=1)
    titulo_ie.value = "INFORMAÇÕES DO ENSAIO"
    titulo_ie.font = Font(name='Arial', size=12, bold=True)
    titulo_ie.alignment = alinhamento_centro
    titulo_ie.fill = fundo_cinza
    for col in range(1, 3):
        ws1.cell(row=linha_ie, column=col).border = borda_fina

    info_labels = [
        ("Massa do operador [kg]",             vehicle_info.get("mass_driver", 0)),
        ("Massa do equipamento [kg]",          vehicle_info.get("mass_equip", 0)),
        ("Massa do veículo [kg]",              vehicle_info.get("mass_vehicle", 0)),
        ("Massa equiv. inércia rotação [kg]",  vehicle_info.get("mass_rot", 0)),
        ("Massa efetiva do veículo [kg]",      vehicle_info.get("mass_total", 0)),
    ]

    for i, (label, val) in enumerate(info_labels):
        row = linha_ie + 1 + i
        ws1.cell(row=row, column=1, value=label).font = fonte_negrito if i == 4 else fonte_padrao

        valor_formatado = round(val, 1) if isinstance(val, (int, float)) else val
        cell_val = ws1.cell(row=row, column=2, value=valor_formatado)
        cell_val.font = fonte_negrito if i == 4 else fonte_padrao
        cell_val.alignment = alinhamento_centro

        for col in range(1, 3):
            ws1.cell(row=row, column=col).border = borda_fina
            

    # ===== ABA 2: TEMPOS E VELOCIDADES =====
    ws2 = wb.create_sheet(title="Tempos e Velocidades")

    # Determinar velocidades a partir dos dados
    velocidades = []
    if all_run_data:
        first_run_key = next(iter(all_run_data))
        if len(all_run_data[first_run_key]) > 1:
            velocidades = all_run_data[first_run_key]["velocities"]

    if not velocidades:
        velocidades = [100 - 5 * i for i in range(16)]

    col_inicio_vel = 3
    col_fim_vel = col_inicio_vel + len(velocidades) - 1
    col_cond_amb = col_fim_vel + 2  # começa bloco de condições ambientais (temp/press/vento)

    # ----- Cabeçalho "Runs" -----
    ws2.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    cell = ws2.cell(row=1, column=2)
    cell.value = "Runs"
    cell.font = fonte_negrito
    cell.alignment = alinhamento_centro
    cell.fill = fundo_cinza
    for rr in (1, 2):
        cell = ws2.cell(row=rr, column=2)
        cell.border = borda_fina

    # ----- Cabeçalho "Velocities [km/h]" -----
    ws2.merge_cells(start_row=1, start_column=col_inicio_vel, end_row=1, end_column=col_fim_vel)
    for col in range(col_inicio_vel, col_fim_vel + 1):
        cell = ws2.cell(row=1, column=col)
        if col == col_inicio_vel:
            cell.value = "Velocities [km/h]"
        cell.font = fonte_negrito
        cell.alignment = alinhamento_centro
        cell.fill = fundo_cinza
        cell.border = borda_fina

    for i, vel in enumerate(velocidades):
        col = col_inicio_vel + i
        cell = ws2.cell(row=2, column=col)
        cell.value = vel
        cell.font = fonte_negrito
        cell.alignment = alinhamento_centro
        cell.border = borda_fina

    # ----- Cabeçalho "Condições Ambientais" (Temp / Press / Vento) -----
    ws2.merge_cells(
        start_row=1,
        start_column=col_cond_amb,
        end_row=1,
        end_column=col_cond_amb + 2
    )
    cond_cell = ws2.cell(row=1, column=col_cond_amb)
    cond_cell.value = "Condições Ambientais"
    cond_cell.font = fonte_negrito
    cond_cell.alignment = alinhamento_centro
    cond_cell.fill = fundo_cinza
    for col in (col_cond_amb, col_cond_amb + 1, col_cond_amb + 2):
        cell = ws2.cell(row=1, column=col)
        cell.border = borda_fina

    ws2.cell(row=2, column=col_cond_amb,     value="Temperatura [°C]").font = fonte_negrito
    ws2.cell(row=2, column=col_cond_amb + 1, value="Pressão [kPa]").font = fonte_negrito
    ws2.cell(row=2, column=col_cond_amb + 2, value="Vento [m/s]").font = fonte_negrito
    for col in (col_cond_amb, col_cond_amb + 1, col_cond_amb + 2):
        cell = ws2.cell(row=2, column=col)
        cell.alignment = alinhamento_centro
        cell.border = borda_fina

    # ----- Dados por PAR (2 linhas: ida / volta) -----
    linha = 3
    for idx, pair_info in enumerate(selected_pairs_details, 1):
        pair_id = pair_info["pair_id"]
        run_ida_str, run_volta_str = pair_id.split("/")
        run_ida = int(run_ida_str)
        run_volta = int(run_volta_str)

        # Tempos por run
        tempos_ida = []
        tempos_volta = []

        if run_ida in all_run_data and len(all_run_data[run_ida]) > 0:
            tempos_ida = all_run_data[run_ida]["times"]
        if run_volta in all_run_data and len(all_run_data[run_volta]) > 0:
            tempos_volta = all_run_data[run_volta]["times"]

        if len(tempos_ida) < len(velocidades):
            tempos_ida.extend([0] * (len(velocidades) - len(tempos_ida)))
        if len(tempos_volta) < len(velocidades):
            tempos_volta.extend([0] * (len(velocidades) - len(tempos_volta)))

        # Coluna PAR (mescla ida/volta)
        ws2.merge_cells(start_row=linha, start_column=1, end_row=linha + 1, end_column=1)
        cell = ws2.cell(row=linha, column=1)
        cell.value = f"Par {idx}"
        cell.font = fonte_negrito
        cell.alignment = alinhamento_centro
        cell.fill = fundo_cinza
        for rr in (linha, linha + 1):
            ws2.cell(row=rr, column=1).border = borda_fina

        # Runs (ida / volta)
        ws2.cell(row=linha,     column=2, value=f"{run_ida_str} ↑ [+]").font = fonte_negrito
        ws2.cell(row=linha + 1, column=2, value=f"{run_volta_str} ↓ [-]").font = fonte_negrito
        for rr in (linha, linha + 1):
            cell = ws2.cell(row=rr, column=2)
            cell.alignment = alinhamento_centro
            cell.border = borda_fina

        # Tempos
        for i, vel in enumerate(velocidades):
            col = col_inicio_vel + i
            t1 = tempos_ida[i] if i < len(tempos_ida) else 0
            t2 = tempos_volta[i] if i < len(tempos_volta) else 0

            for row_val, val in [(linha, t1), (linha + 1, t2)]:
                cell = ws2.cell(row=row_val, column=col)
                cell.value = round(val, 2) if isinstance(val, (int, float)) else 0
                cell.font = fonte_padrao
                cell.alignment = alinhamento_centro
                cell.border = borda_fina

        # ----- Condições ambientais individuais por passada -----
        temp_ida   = pair_info.get("temp_ida_used",  pair_info.get("temp", None))
        temp_volta = pair_info.get("temp_volta_used", pair_info.get("temp", None))
        press_ida  = pair_info.get("press_ida_used", pair_info.get("press", None))
        press_volta= pair_info.get("press_volta_used", pair_info.get("press", None))
        wind_ida   = pair_info.get("wind_ida_ms", None)
        wind_volta = pair_info.get("wind_volta_ms", None)

        # Linha ida
        cT_ida = ws2.cell(row=linha, column=col_cond_amb)
        cP_ida = ws2.cell(row=linha, column=col_cond_amb + 1)
        cW_ida = ws2.cell(row=linha, column=col_cond_amb + 2)

        cT_ida.value = temp_ida if isinstance(temp_ida, (int, float)) else ""
        cP_ida.value = press_ida if isinstance(press_ida, (int, float)) else ""
        cW_ida.value = wind_ida if isinstance(wind_ida, (int, float)) else ""

        for c in (cT_ida, cP_ida, cW_ida):
            c.font = fonte_padrao
            c.alignment = alinhamento_centro
            c.border = borda_fina

        # Linha volta
        cT_volta = ws2.cell(row=linha + 1, column=col_cond_amb)
        cP_volta = ws2.cell(row=linha + 1, column=col_cond_amb + 1)
        cW_volta = ws2.cell(row=linha + 1, column=col_cond_amb + 2)

        cT_volta.value = temp_volta if isinstance(temp_volta, (int, float)) else ""
        cP_volta.value = press_volta if isinstance(press_volta, (int, float)) else ""
        cW_volta.value = wind_volta if isinstance(wind_volta, (int, float)) else ""

        for c in (cT_volta, cP_volta, cW_volta):
            c.font = fonte_padrao
            c.alignment = alinhamento_centro
            c.border = borda_fina

        linha += 2



    # ===== ABA 3: COEFICIENTES CALCULADOS =====
    ws3 = wb.create_sheet(title="Coeficientes Calculados")

    # Cabeçalhos
    headers_ws3 = [
        "Par",
        "Run",
        "f'0 (bruto)",
        "f'2 (bruto)",
        "Média f'0 (bruto)",
        "Média f'2 (bruto)",
    ]
    for col, header in enumerate(headers_ws3, start=1):
        cell = ws3.cell(row=1, column=col)
        cell.value = header
        cell.font = fonte_negrito
        cell.fill = fundo_cinza
        cell.alignment = alinhamento_centro
        cell.border = borda_fina

    row_idx = 2  # começa depois do cabeçalho

    for idx, pair_info in enumerate(selected_pairs_details, start=1):
        # Runs (podem vir como run1/run2 ou run_ida/run_volta)
        run_ida = pair_info.get("run1", pair_info.get("run_ida", "N/A"))
        run_volta = pair_info.get("run2", pair_info.get("run_volta", "N/A"))

        # Coeficientes brutos por passada
        # tenta pegar *_raw; se não existir, cai em f0_ida/f0_volta para não quebrar
        f0_ida_raw   = pair_info.get("f0_ida_raw",   pair_info.get("f0_ida"))
        f2_ida_raw   = pair_info.get("f2_ida_raw",   pair_info.get("f2_ida"))
        f0_volta_raw = pair_info.get("f0_volta_raw", pair_info.get("f0_volta"))
        f2_volta_raw = pair_info.get("f2_volta_raw", pair_info.get("f2_volta"))

        vals_f0_raw = [v for v in (f0_ida_raw, f0_volta_raw) if isinstance(v, (int, float))]
        vals_f2_raw = [v for v in (f2_ida_raw, f2_volta_raw) if isinstance(v, (int, float))]
        mean_f0_raw = statistics.mean(vals_f0_raw) if vals_f0_raw else None
        mean_f2_raw = statistics.mean(vals_f2_raw) if vals_f2_raw else None

        # ----- Linha da IDA -----
        c = ws3.cell(row=row_idx, column=1)
        c.value = f"Par {idx}"
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina

        c = ws3.cell(row=row_idx, column=2)
        c.value = f"{run_ida} ↑ [+]"
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina

        c = ws3.cell(row=row_idx, column=3)
        c.value = f0_ida_raw if isinstance(f0_ida_raw, (int, float)) else None
        c.font = fonte_padrao
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.0000"

        c = ws3.cell(row=row_idx, column=4)
        c.value = f2_ida_raw if isinstance(f2_ida_raw, (int, float)) else None
        c.font = fonte_padrao
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.000000"

        c = ws3.cell(row=row_idx, column=5)
        c.value = mean_f0_raw if isinstance(mean_f0_raw, (int, float)) else None
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.0000"

        c = ws3.cell(row=row_idx, column=6)
        c.value = mean_f2_raw if isinstance(mean_f2_raw, (int, float)) else None
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.000000"

        # ----- Linha da VOLTA -----
        c = ws3.cell(row=row_idx + 1, column=1)
        c.value = ""  # deixa em branco, o "Par X" só na primeira
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina

        c = ws3.cell(row=row_idx + 1, column=2)
        c.value = f"{run_volta} ↓ [-]"
        c.font = fonte_negrito
        c.alignment = alinhamento_centro
        c.border = borda_fina

        c = ws3.cell(row=row_idx + 1, column=3)
        c.value = f0_volta_raw if isinstance(f0_volta_raw, (int, float)) else None
        c.font = fonte_padrao
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.0000"

        c = ws3.cell(row=row_idx + 1, column=4)
        c.value = f2_volta_raw if isinstance(f2_volta_raw, (int, float)) else None
        c.font = fonte_padrao
        c.alignment = alinhamento_centro
        c.border = borda_fina
        c.number_format = "0.000000"

        # colunas de médias na linha de baixo: só borda / alinhamento
        for col in (5, 6):
            c = ws3.cell(row=row_idx + 1, column=col)
            c.border = borda_fina
            c.alignment = alinhamento_centro
            if col == 5:
                c.number_format = "0.0000"
            else:
                c.number_format = "0.000000"

        row_idx += 2

    # Salvar arquivo
    full_path = os.path.join(output_path, filename)
    wb.save(full_path)
    return full_path
