# coding: utf-8
"""Pure Excel export for final Split results."""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime
import io
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.split_deviation_analysis import (
    analyze_split_selected_deviations,
    build_selected_pairs_signature,
)
from core.split_display import get_split_pair_public_label
from core.split_results import consolidate_split_final_results
from core.split_weather_context import split_environmental_values
from core.split_vehicle_mass import normalize_split_vehicle_mass_data


COMPONENTS = ("high_plus", "low_plus", "high_minus", "low_minus")
COMPONENT_LABELS = {
    "high_plus": "High [+]",
    "low_plus": "Low [+]",
    "high_minus": "High [-]",
    "low_minus": "Low [-]",
}
MISSING = "-"
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7") for _ in range(4)))
STATUS_FILLS = {
    "approved": PatternFill("solid", fgColor="0EE427"),
    "warning": PatternFill("solid", fgColor="E6F200"),
    "failed": PatternFill("solid", fgColor="FF5757"),
}


def _freeze_signature_value(value):
    if isinstance(value, dict):
        return tuple(
            (str(key), _freeze_signature_value(item))
            for key, item in sorted(value.items(), key=lambda entry: str(entry[0]))
        )
    if isinstance(value, (list, tuple, set)):
        return tuple(_freeze_signature_value(item) for item in value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def build_split_export_signature(
    *, final_results: dict, selected_pairs: list[dict],
    vehicle_data: dict, deviation_analysis: dict,
) -> tuple:
    """Return a stable signature for the complete Split workbook inputs."""
    return (
        build_selected_pairs_signature(selected_pairs),
        _freeze_signature_value(final_results or {}),
        _freeze_signature_value(vehicle_data or {}),
        _freeze_signature_value(deviation_analysis or {}),
    )


def get_cached_split_export(
    signature: tuple, cache: dict | None, *, builder,
) -> tuple[bytes, dict, bool]:
    """Build workbook bytes only when the supplied export signature changed."""
    current = cache if isinstance(cache, dict) else {}
    if current.get("signature") == signature and isinstance(current.get("payload"), bytes):
        return current["payload"], current, True
    payload = builder()
    refreshed = {"signature": signature, "payload": payload}
    return payload, refreshed, False


def _number(value):
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _value(value):
    if value is None:
        return MISSING
    if isinstance(value, float) and not math.isfinite(value):
        return MISSING
    if isinstance(value, (list, tuple, set)):
        text = "; ".join(str(item) for item in value if str(item).strip())
        return text or MISSING
    if isinstance(value, dict):
        text = "; ".join(f"{key}: {_value(item)}" for key, item in value.items())
        return text or MISSING
    text = str(value).strip()
    return text if text and text.lower() not in {"nan", "none", "n/a"} else MISSING


def _status(value):
    if value is True:
        return "Aprovado"
    if value is False:
        return "Reprovado"
    if value is None:
        return "Inconclusivo"
    return {
        "conforming": "Aprovado",
        "approved": "Aprovado",
        "nonconforming": "Reprovado",
        "failed": "Reprovado",
        "not_evaluable": "Inconclusivo",
        "insufficient_data": "Inconclusivo",
        "incomplete": "Incompleto",
        "warning": "Atenção",
    }.get(value, _value(value))


def _selected_pairs(pairs):
    return [deepcopy(pair) for pair in (pairs or []) if isinstance(pair, dict) and pair.get("selected") is True]


def _component(pair, component):
    record = pair.get(component)
    return record if isinstance(record, dict) else {}


def _component_label(pair, component):
    record = _component(pair, component)
    run = record.get("run_id", pair.get(f"{component}_run"))
    delta_t = _number(record.get("delta_t_s", pair.get(f"{component}_delta_t_s")))
    if run in (None, ""):
        return MISSING
    return f"Run {run}" + (f" | dt = {delta_t:.3f} s" if delta_t is not None else "")


def _origin_label(pair):
    sources = pair.get("algorithm_sources") or []
    if isinstance(sources, str):
        sources = [sources]
    normalized = {str(value).strip().lower() for value in sources}
    algorithm_source = str(pair.get("algorithm_source") or "").strip().lower()
    if algorithm_source:
        normalized.add(algorithm_source)
    if pair.get("selected_by_energy_algo"):
        normalized.add("energy")
    if pair.get("selected_by_target_algo"):
        normalized.add("target")
    if {"energy", "target"}.issubset(normalized):
        return "Energia + Target"
    if "energy" in normalized:
        return "Energia"
    if "target" in normalized:
        return "Target"
    if str(pair.get("selection_source") or "").strip().lower() == "manual":
        return "Manual"
    return "Desconhecido"


def _weather_for_pair(pair):
    environmental = split_environmental_values(pair)
    temperature = environmental["temperature_c"]
    pressure = environmental["pressure_kpa"]
    wind = environmental["wind_speed_mps"]
    alerts = []
    if wind is not None and wind > 3.0:
        alerts.append("Vento acima de 3 m/s")
    if temperature is not None and temperature > 35.0:
        alerts.append("Temperatura acima de 35 °C")
    return temperature, pressure, wind, alerts


def _append_table(ws, start_row, title, headers, rows):
    ws.cell(start_row, 1, title).font = Font(bold=True, size=12)
    ws.cell(start_row, 1).fill = TITLE_FILL
    header_row = start_row + 1
    for column, header in enumerate(headers, 1):
        cell = ws.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for row_index, row in enumerate(rows, header_row + 1):
        for column, item in enumerate(row, 1):
            numeric = _number(item)
            cell_value = numeric if numeric is not None else _value(item)
            cell = ws.cell(row_index, column, cell_value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return header_row + len(rows) + 2


def _finish_sheet(ws):
    ws.freeze_panes = None
    for column in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1)) + 2
        ws.column_dimensions[get_column_letter(column)].width = min(max(width, 12), 42)
    ws.auto_filter.ref = None


def _status_fill_key(value):
    normalized = str(value or "").strip().casefold()
    if normalized in {"approved", "aprovado", "ok", "valid", "válido", "valido"}:
        return "approved"
    if normalized in {"failed", "reprovado", "fora do limite", "invalid", "inválido", "invalido"}:
        return "failed"
    if normalized in {"warning", "atenção", "atencao", "inconclusivo"}:
        return "warning"
    return None


def _merge_summary_title(ws, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )


def _vehicle_rows(vehicle_data):
    source = deepcopy(vehicle_data or {})
    nested = source.pop("vehicle_info", None)
    if isinstance(nested, dict):
        source = {**nested, **source}
    mass_data = normalize_split_vehicle_mass_data(vehicle_data)
    rows = [
        ("Modelo do veículo", source.get("model")),
        ("Data do ensaio", source.get("test_date")),
        ("Massa em ordem de marcha [kg]", mass_data["running_order_mass_kg"]),
        ("Massa de ensaio M [kg]", mass_data["test_mass_kg"]),
        ("Massa equivalente de rotação me [kg]", mass_data["rotational_equivalent_mass_kg"]),
        ("Massa efetiva Me [kg]", mass_data["effective_mass_kg"]),
    ]
    return [(label, value) for label, value in rows if value is not None]


def _weather_summary(pairs):
    values = [split_environmental_values(pair) for pair in pairs]
    temperatures = [item["temperature_c"] for item in values if item["temperature_c"] is not None]
    pressures = [item["pressure_kpa"] for item in values if item["pressure_kpa"] is not None]
    winds = [item["wind_speed_mps"] for item in values if item["wind_speed_mps"] is not None]
    modes = {str(item.get("mode") or "").lower() for item in values}
    environment = (
        "Parâmetros fixos" if modes and modes <= {"fixed"}
        else "Meteorologia sincronizada" if values and modes - {"fixed", ""}
        else MISSING
    )
    alerts = list(dict.fromkeys(
        alert for pair in pairs for alert in _weather_for_pair(pair)[3]
    ))
    return (
        environment,
        sum(temperatures) / len(temperatures) if temperatures else None,
        sum(pressures) / len(pressures) if pressures else None,
        max(winds) if winds else None,
        alerts,
    )


def _write_summary(wb, final_results, generated_at, vehicle_data, pairs):
    ws = wb.active
    ws.title = "Resumo Final"
    title_rows = [1]
    row = _append_table(ws, 1, "RESULTADOS SPLIT", ("Campo", "Valor"), [
        ("Método", "Split"),
        ("Data de geração", generated_at.strftime("%d/%m/%Y %H:%M:%S")),
        ("Quantidade de pares selecionados", final_results.get("num_pairs")),
        ("Status final", _status(final_results.get("conformity_status"))),
    ])
    title_rows.append(row)
    row = _append_table(ws, row, "RESULTADO FINAL", ("Campo", "Valor"), [
        ("F0 final [N]", final_results.get("mean_f0")),
        ("F2 final [N/(km/h)²]", final_results.get("mean_f2")),
        ("CV F0 [%]", final_results.get("cv_f0")),
        ("CV F2 [%]", final_results.get("cv_f2")),
        ("Energia média [MJ/km]", final_results.get("mean_energy")),
    ])
    title_rows.append(row)
    row = _append_table(ws, row, "DADOS DO VEÍCULO", ("Campo", "Valor"), _vehicle_rows(vehicle_data) or [(MISSING, MISSING)])
    title_rows.append(row)
    environment, temperature, pressure, wind, alerts = _weather_summary(pairs)
    _append_table(ws, row, "RESUMO METEOROLÓGICO", ("Campo", "Valor"), [
        ("Ambiente", environment),
        ("Temperatura média [°C]", temperature),
        ("Pressão média [kPa]", pressure),
        ("Vento máximo [m/s]", wind),
        ("Alertas meteorológicos", alerts),
    ])
    formats = {
        "F0 final [N]": "0.0000", "F2 final [N/(km/h)²]": "0.000000",
        "CV F0 [%]": "0.00", "CV F2 [%]": "0.00", "Energia média [MJ/km]": "0.0000",
    }
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row_index, 1).value in formats:
            ws.cell(row_index, 2).number_format = formats[ws.cell(row_index, 1).value]
        if ws.cell(row_index, 1).value == "Status final":
            status_cell = ws.cell(row_index, 2)
            fill_key = _status_fill_key(status_cell.value)
            if fill_key:
                status_cell.fill = STATUS_FILLS[fill_key]
                status_cell.font = Font(bold=True, color="000000")
    for title_row in title_rows:
        _merge_summary_title(ws, title_row)
    _finish_sheet(ws)


def _write_pairs(wb, pairs):
    ws = wb.create_sheet("Pares Selecionados")
    headers = ("Par", "Origem", "High [+]", "Low [+]", "High [-]", "Low [-]", "DeltaT high+ [s]", "DeltaT low+ [s]", "DeltaT high- [s]", "DeltaT low- [s]", "F0 [N]", "F2 [N/(km/h)²]", "CV F0 [%]", "CV F2 [%]", "Energia [MJ/km]", "Temperatura [°C]", "Pressão [kPa]", "Vento [m/s]", "Alertas meteorológicos")
    rows = []
    for pair in pairs:
        temperature, pressure, wind, alerts = _weather_for_pair(pair)
        rows.append((get_split_pair_public_label(pair), _origin_label(pair), *(_component_label(pair, component) for component in COMPONENTS), *(pair.get(f"{component}_delta_t_s") if pair.get(f"{component}_delta_t_s") is not None else _component(pair, component).get("delta_t_s") for component in COMPONENTS), pair.get("F0_mean", pair.get("F0")), pair.get("F2_mean", pair.get("F2")), pair.get("cv_F0_percent"), pair.get("cv_F2_percent"), pair.get("energy"), temperature, pressure, wind, alerts))
    _append_table(ws, 1, "PARES USADOS NO RESULTADO FINAL", headers, rows)
    for row in range(3, ws.max_row + 1):
        for column, number_format in ((11, "0.0000"), (12, "0.000000"), (13, "0.00"), (14, "0.00"), (15, "0.0000")):
            ws.cell(row, column).number_format = number_format
    _finish_sheet(ws)


def _write_deviations(wb, analysis, pairs):
    ws = wb.create_sheet("Análise de Desvios e Tempos")
    coefficients = analysis.get("coefficient_summary") or {}
    row = _append_table(ws, 1, "RESUMO CV F0/F2", ("Coeficiente", "Média", "CV [%]", "Limite [%]", "Status"), [
        ("F0", coefficients.get("mean_f0"), coefficients.get("cv_f0_pct"), coefficients.get("limit_pct"), _status(coefficients.get("status"))),
        ("F2", coefficients.get("mean_f2"), coefficients.get("cv_f2_pct"), coefficients.get("limit_pct"), _status(coefficients.get("status"))),
    ])
    deviations = analysis.get("pair_deviations") or []
    deviation_rows = []
    for index, item in enumerate(deviations):
        label = get_split_pair_public_label(pairs[index]) if index < len(pairs) else MISSING
        deviation_rows.append((label, item.get("f0"), item.get("f0_deviation_pct"), item.get("f2"), item.get("f2_deviation_pct"), item.get("energy"), item.get("alerts")))
    row = _append_table(ws, row, "DESVIOS POR PAR", ("Par", "F0 [N]", "Desvio F0 [%]", "F2 [N/(km/h)²]", "Desvio F2 [%]", "Energia [MJ/km]", "Alerta"), deviation_rows)
    times = analysis.get("time_summary") or {}
    groups = times.get("groups") or {}
    row = _append_table(ws, row, "TEMPOS DELTAT", ("Grupo", "n", "Média deltaT [s]", "Desvio padrão [s]", "CV deltaT [%]", "Limite [%]", "Status"), [(COMPONENT_LABELS.get(key, key), value.get("count"), value.get("mean"), value.get("stdev"), value.get("cv_pct"), times.get("cv_limit_pct"), _status(value.get("passed"))) for key, value in groups.items()])
    opposite = times.get("opposite_direction") or {}
    _append_table(ws, row, "DIFERENÇA ENTRE SENTIDOS", ("Velocidade", "Média [+] [s]", "Média [-] [s]", "Diferença [%]", "Limite [%]", "Status"), [(key.title(), value.get("mean_plus"), value.get("mean_minus"), value.get("diff_pct"), times.get("opposite_mean_limit_pct"), _status(value.get("passed"))) for key, value in opposite.items()])
    _finish_sheet(ws)


def export_split_final_results_to_excel(*, final_results: dict | None, selected_pairs: list[dict], vehicle_data: dict, deviation_analysis: dict | None = None, generated_at: datetime | None = None) -> bytes:
    """Return an .xlsx report built only from explicitly selected Split pairs."""
    pairs = _selected_pairs(selected_pairs)
    consolidated = consolidate_split_final_results(pairs)
    if final_results:
        # The core consolidation remains authoritative; retain extra warnings only.
        consolidated["warnings"] = list(dict.fromkeys([*(consolidated.get("warnings") or []), *((final_results or {}).get("warnings") or [])]))
    analysis = deepcopy(deviation_analysis) if deviation_analysis is not None else analyze_split_selected_deviations(pairs)
    wb = Workbook()
    _write_summary(wb, consolidated, generated_at or datetime.now(), vehicle_data, pairs)
    _write_pairs(wb, pairs)
    _write_deviations(wb, analysis, pairs)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
